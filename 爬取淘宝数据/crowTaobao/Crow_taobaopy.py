#!/usr/bin/env python
# coding: utf-8

# In[1]:


import requests
import csv
import codecs
import re
import urllib.request
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


# In[2]:


def get_html_text(url):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
    try:
        coo = 't=85db5e7cb0133f23f29f98c7d6955615; cna=3uklFEhvXUoCAd9H6ovaVLTG; isg=BM3NGT0Oqmp6Mg4qfcGPnvDY3-pNqzF2joji8w9SGWTYBu241_taTS6UdFrF3Rk0; miid=983575671563913813; thw=cn; um=535523100CBE37C36EEFF761CFAC96BC4CD04CD48E6631C3112393F438E181DF6B34171FDA66B2C2CD43AD3E795C914C34A100CE538767508DAD6914FD9E61CE; _cc_=W5iHLLyFfA%3D%3D; tg=0; enc=oRI1V9aX5p%2BnPbULesXvnR%2BUwIh9CHIuErw0qljnmbKe0Ecu1Gxwa4C4%2FzONeGVH9StU4Isw64KTx9EHQEhI2g%3D%3D; hng=CN%7Czh-CN%7CCNY%7C156; mt=ci=0_0; hibext_instdsigdipv2=1; JSESSIONID=EC33B48CDDBA7F11577AA9FEB44F0DF3'
        cookies = {}
        for line in coo.split(';'):  # 浏览器伪装
            name, value = line.strip().split('=', 1)
            cookies[name] = value
        r = requests.get(url, cookies=cookies, headers=headers, timeout=30)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return ''


def parse_page(ilt_list, html):
    #items = doc('#mainsrp-itemlist .items').items()
    #print(type(items))
    plt = re.findall(r'\"view_price\"\:\"[\d\.]*\"', html)
    tlt = re.findall(r'\"raw_title\"\:\".*?\"', html)
    shlt = re.findall(r'\"nick\"\:\".*?\"', html)
    urlt = re.findall(r'\"detail_url\"\:\".*?\"',html)
    pic_urlt = re.findall(r'\"pic_url\"\:\".*?\"',html)
    for i in range(len(plt)):
        price = eval(plt[i].split(':')[1])
        title = eval(tlt[i].split(':')[1])
        shop_name = eval(shlt[i].split(':')[1])
        url = urlt[i].split(':')[1]
        pic_url = pic_urlt[i].split(':')[1]
        ilt_list.append([price,title,shop_name,url,pic_url])
    return ilt_list

def data_write_csv(file_name, datas):#file_name为写入CSV文件的路径，datas为要写入数据列表
    file_csv = codecs.open(file_name,'w+','utf-8')#追加
    writer = csv.writer(file_csv, delimiter=' ', quotechar=' ', quoting=csv.QUOTE_MINIMAL)
    for data in datas:
        writer.writerow(data)
    print("保存文件成功，处理结束")

def SaveToExcel(lit):
    list = pd.DataFrame(data = lit)
    list.rename(columns={0:'价格',1:'名称',2:'店名',3:'商品URL',4:'缩略图URL'},inplace=True)
    list['电商平台'] = '淘宝'
    list['关键字'] = '军帽'
    list['缩略图'] = '图'
    list.to_excel('./1.xlsx',encoding = 'utf-8')
    
def DownloadPic(lit):
    pic_url = [i[4] for i in lit]
    i = 0
    for img_url in pic_url:
        filepath = 'D:\Privatefiles\Jupyter notebook\job_three\imgs' + str(i) + '.jpg'
        print(filepath)
        i = i + 1
        f = open(filepath,'wb')
        #req = urllib.request.urlopen('http:' + img_url)
        req = urllib.request.urlretrieve(img_url,filename = filepath)
        print(req)
        buf = req.read()
        f.write(buf)
 
def write_tofile(img_urls,page):
    for id,img_url in enumerate(img_urls):
        pic = requests.get(img_url)
        with open('/{}_{}.jpg'.format(page,id),'wb') as f:
            f.write(pic.content)
        
def get_pic_by_url(folder_path, lists):
    pic_url = [i[4] for i in lists]  
    if not os.path.exists(folder_path):
        print(1)
        print("Selected folder not exist, try to create it.")
        os.makedirs(folder_path)
    i = 0
    for url in pic_url:
        print("Try downloading file: {}".format(url))
        filepath = folder_path + '/' + str(i) + '.jpg'
        i = i + 1
        print(filepath)
        if os.path.exists(filepath):
            print("File have already exist. skip")
        else:
            try:
                url = "https:" + eval(url)
                urllib.request.urlretrieve(url, filename = filepath)
                print("successful")
            except Exception as e:
                print("Error occurred when downloading file, error message:")
                print(e)

def InsertImg(imgPath, excelPath, insertPath):
    imgsize = (720 / 12, 1280 / 20)  # 设置一个图像缩小的比例
    wb = load_workbook(excelPath)
    ws = wb.active  
    width = 10.0
    height = 8 * (2.2862 / 0.3612)
    for i in range(1, ws.max_row+1):
        ws.row_dimensions[i].height = height
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].width = width   
    img = Image(imgPath)  # 缩放图片
    img.width, img.height = imgsize
    ws.add_image(img, insertPath)  # 图片 插入 A1 的位置上
    wb.save('1.xlsx')  # 新的结果保存输出

def InsertAllImg():
    i = 0
    for i in range(224):
        imgPath = './imgs/' + str(i) + '.jpg'
        print(imgPath)
        insertPath = 'I' + str(i+2)
        print(insertPath)
        InsertImg(imgPath, './1.xlsx', insertPath)

def main():
    keyword = '军帽' #指定要爬取的商品
    depth = 5  # 要爬取几页
    start_url = 'https://s.taobao.com/search?q=' + keyword
    info_list= []
    for i in range(depth):
        try:
            url = start_url + '&s=' + str(44 * i)  # 44是淘宝每个页面呈现的宝贝数量
            html = get_html_text(url)
            #print(html)
            parse_page(info_list, html)
            #write_product(info_dic)
        except:
            print("wrong")
    SaveToExcel(info_list)
    #DownloadPic(info_list)
    #path = "./imgs"
    #get_pic_by_url(path, info_list)
    InsertAllImg()
main()
